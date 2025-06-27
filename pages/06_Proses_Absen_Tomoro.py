import pandas as pd
import streamlit as st
import io
from datetime import datetime, time, date, timedelta
import re
import math
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
# Removed import for PAPERSIZE, ORIENTATION as per user request to not hardcode A4 landscape

# --- Inisialisasi variabel ---
df_processed = None
bulan_laporan_val = None
tahun_laporan_val = None

# Konfigurasi halaman
st.set_page_config(layout="wide", page_title="Dashboard Tools Ni Wayan Astari")
st.title("Convert Log Absensi Mesin Eppos")
st.write("Unggah file Excel laporan sidik jari Kamu di sini untuk diproses.")

# Helper function to convert HH:MM string to total minutes
def time_to_minutes(time_str):
    if pd.isna(time_str) or not isinstance(time_str, str) or ':' not in time_str:
        return 0
    try:
        hours, minutes = map(int, time_str.split(':'))
        return hours * 60 + minutes
    except ValueError:
        return 0

# Helper function to convert total minutes back to HH:MM string
def minutes_to_hhmm(total_minutes):
    if total_minutes is None:
        return ""
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{int(hours):02}:{int(minutes):02}"

def process_attendance_log(uploaded_file):
    st.info("Memulai pengolahan data log karyawan dari laporan sidik jari...")

    try:
        df_raw = pd.read_excel(io.BytesIO(uploaded_file.read()), header=None)
        st.success(f"File '{uploaded_file.name}' berhasil dibaca.")
    except Exception as e:
        st.error(f"Gagal membaca file Excel: {e}")
        return None, None, None

    if df_raw.empty:
        st.warning("Sheet yang dibaca kosong.")
        return None, None, None

    st.subheader("DEBUG INFO: 10 Baris Pertama dari Sheet Mentah")
    st.dataframe(df_raw.head(10))

    start_date_full = None
    end_date_full = None
    periode_row_index = -1 

    # Find the row containing the "Periode" to extract full start and end dates
    for r_idx, row_series in df_raw.iterrows():
        row_str = ' '.join(row_series.dropna().astype(str).tolist())
        # Regex to capture YYYY/MM/DD (start date) and MM/DD (end date)
        match = re.search(r'Periode\s*:\s*(\d{4})/(\d{2})/(\d{2})\s*~\s*(\d{2})/(\d{2})', row_str)
        if match:
            start_year = int(match.group(1))
            start_month = int(match.group(2))
            start_day = int(match.group(3))
            end_month_parsed = int(match.group(4))
            end_day_parsed = int(match.group(5))

            start_date_full = date(start_year, start_month, start_day)
            
            # Determine the correct year for the end date
            end_year_full = start_year
            # If the end month is numerically smaller than the start month, it means it's in the next year
            if end_month_parsed < start_month: 
                end_year_full += 1
            # If months are the same but end day is smaller, it could be a period like Dec 20 - Jan 19 (next year) or May 20 - May 19 (next year)
            elif end_month_parsed == start_month and end_day_parsed < start_day:
                end_year_full += 1
            
            end_date_full = date(end_year_full, end_month_parsed, end_day_parsed)

            periode_row_index = r_idx
            break

    if start_date_full is None or end_date_full is None:
        st.error("Periode tidak ditemukan dalam file. Pastikan format 'Periode :YYYY/MM/DD ~ DD/MM' ada.")
        return None, None, None

    # Generate all actual dates in the period (e.g., May 26, May 27, ..., June 25)
    all_period_dates = []
    current_date_iter = start_date_full
    while current_date_iter <= end_date_full:
        all_period_dates.append(current_date_iter)
        current_date_iter += timedelta(days=1)

    days_row_idx_global = periode_row_index + 1
    col_date_mapping_global = {} # Maps actual date objects to their column index

    # Populate col_date_mapping_global by matching header day numbers with the chronological dates
    date_idx_in_period = 0
    for c_idx in range(df_raw.shape[1]):
        if date_idx_in_period >= len(all_period_dates): # Stop if all dates in our sequence have been mapped
            break
        try:
            cell_val_day = df_raw.iloc[days_row_idx_global, c_idx]
            if pd.isna(cell_val_day):
                continue
            day_num_from_header = int(float(cell_val_day))
            
            expected_date = all_period_dates[date_idx_in_period]
            
            # Check if the day number from the header matches the day of the expected date
            if day_num_from_header == expected_date.day:
                col_date_mapping_global[expected_date] = c_idx
                date_idx_in_period += 1 # Move to the next expected date
            # If it doesn't match, it might be a blank column or a non-day value; skip this column for date mapping.
        except (ValueError, TypeError):
            continue

    if not col_date_mapping_global or len(col_date_mapping_global) != len(all_period_dates):
        st.error(f"Nomor hari di header tidak sepenuhnya cocok dengan periode yang ditemukan. Diharapkan {len(all_period_dates)} hari dari periode '{start_date_full.strftime('%Y/%m/%d')} ~ {end_date_full.strftime('%m/%d')}', tetapi {len(col_date_mapping_global)} kolom hari yang cocok ditemukan.")
        return None, None, None

    # Derive bulan_laporan and tahun_laporan from the start date of the period for consistency (though not directly used in date logic anymore)
    bulan_laporan = start_date_full.month
    tahun_laporan = start_date_full.year

    raw_logs = [] # To store all extracted raw attendance logs
    employee_blocks_indices = [] # To store starting row indices of each employee's block

    # Identify the start of each employee's block (rows containing 'No :')
    for r_idx in range(df_raw.shape[0]):
        # Check if 'No :' is present in the first or second column of the row
        if 'No :' in str(df_raw.iloc[r_idx, 0]) or 'No :' in str(df_raw.iloc[r_idx, 1]):
            employee_blocks_indices.append(r_idx)

    if not employee_blocks_indices:
        st.error("Blok karyawan tidak ditemukan. Pastikan setiap karyawan memiliki baris 'No :'.")
        return None, None, None

    # Process each employee's block
    for i, start_row_idx in enumerate(employee_blocks_indices):
        nama_karyawan = None
        name_label_col_idx = -1 # Column index of 'Nama :'

        # Find 'Nama :' label in the current employee block's header row
        for col_idx in range(df_raw.shape[1]):
            cell_val = str(df_raw.iloc[start_row_idx, col_idx]).strip()
            if 'Nama :' in cell_val:
                name_label_col_idx = col_idx
                break
        
        # Extract employee name (it's in the column(s) after 'Nama :')
        if name_label_col_idx != -1:
            for scan_col_idx in range(name_label_col_idx + 1, df_raw.shape[1]):
                candidate_name = str(df_raw.iloc[start_row_idx, scan_col_idx]).strip()
                # A valid name should not be empty, 'nan', or start with 'Dept :'
                if candidate_name and candidate_name != 'nan' and not candidate_name.startswith('Dept :'):
                    nama_karyawan = candidate_name
                    break

        if not nama_karyawan:
            st.warning(f"Nama karyawan tidak ditemukan di blok dimulai dari baris {start_row_idx + 1}. Melewatkan blok ini.")
            continue

        # Determine the end row for the current employee's log entries
        # It's either the start of the next employee's block or the end of the dataframe
        end_row_idx = employee_blocks_indices[i + 1] if i + 1 < len(employee_blocks_indices) else df_raw.shape[0]
        
        logs_start_row_for_this_block = start_row_idx + 1 # Log entries start after the header row

        # Iterate through each actual date in the period to extract attendance times
        # Use sorted keys from col_date_mapping_global to process dates in order
        for current_date in sorted(col_date_mapping_global.keys()):
            col_idx = col_date_mapping_global[current_date]
            
            # Iterate through rows within the current employee's log block for the specific date's column
            for r_idx_log in range(logs_start_row_for_this_block, end_row_idx):
                # Ensure we don't go out of bounds for columns
                if col_idx >= df_raw.shape[1]:
                    continue

                log_cell_value = str(df_raw.iloc[r_idx_log, col_idx]).strip()

                # Skip empty or NaN cells
                if pd.isna(log_cell_value) or log_cell_value == '' or log_cell_value == 'nan':
                    continue
                
                # Split cell content by spaces or newlines to find multiple times
                potential_times = re.split(r'\s+|\n', log_cell_value)

                for time_str in potential_times:
                    time_str = time_str.strip()
                    if not time_str:
                        continue

                    parsed_time = None
                    # Try parsing with %H:%M:%S first, then %H:%M
                    try:
                        parsed_time = datetime.strptime(time_str, '%H:%M:%S').time()
                    except ValueError:
                        try:
                            parsed_time = datetime.strptime(time_str, '%H:%M').time()
                        except ValueError:
                            # If neither format works, skip this string
                            continue

                    if parsed_time:
                        raw_logs.append({
                            'Nama': nama_karyawan,
                            'Tanggal': current_date, # Use the full date object
                            'Jam': parsed_time
                        })

    if not raw_logs:
        st.warning("Tidak ada log absensi yang berhasil diekstrak dari file. Akan mencoba mengisi dengan tanggal kosong.")
        # If no raw logs are found, proceed to create an empty DataFrame structure
        # This allows the subsequent logic to still generate output with blank dates
        # for employees if the file format is valid but data is empty.
        df_log_mentah = pd.DataFrame(columns=['Nama', 'Tanggal', 'Jam'])
    else:
        df_log_mentah = pd.DataFrame(raw_logs)
        # Sort logs by Name, then Date, then Time to ensure correct order for processing
        df_log_mentah.sort_values(by=['Nama', 'Tanggal', 'Jam'], inplace=True)


    # Define final columns based on user's screenshot request
    kolom_final = ['Nama', 'Tanggal', 'Data Log Mentah', 'Jam Datang', 'Jam Pulang', 'Durasi Jam Kerja']

    log_karyawan_harian = {} # Dictionary to aggregate logs per employee per day

    # Get unique employee names from the raw data or employee blocks
    all_employee_names = df_log_mentah['Nama'].unique().tolist() if not df_log_mentah.empty else [
        # Try to extract employee names even if no logs, from the identified blocks
        str(df_raw.iloc[idx, col_idx + 1]).strip() 
        for idx in employee_blocks_indices 
        for col_idx in range(df_raw.shape[1]) 
        if 'Nama :' in str(df_raw.iloc[idx, col_idx]).strip() 
        and str(df_raw.iloc[idx, col_idx + 1]).strip() != 'nan'
    ]
    all_employee_names = sorted(list(set(name for name in all_employee_names if name))) # Remove duplicates and ensure no empty strings

    if not all_employee_names:
        st.error("Tidak ada nama karyawan yang ditemukan dalam file.")
        return None, None, None

    # Pre-populate log_karyawan_harian with all employee-date combinations
    for emp_name in all_employee_names:
        for single_date in all_period_dates: # Iterate through all valid dates in the period
            key = (emp_name, single_date)
            log_karyawan_harian[key] = {
                'Nama': emp_name,
                'Tanggal': single_date,
                'Data Log Mentah': [], # This will be filled with actual times
                'Jam Datang': None,
                'Jam Pulang': None,
                'Durasi Jam Kerja': None,
            }
    
    # Now, populate the 'Data Log Mentah' with actual raw times
    for index, row in df_log_mentah.iterrows():
        nama = row['Nama']
        tanggal = row['Tanggal']
        jam_obj = row['Jam']
        key = (nama, tanggal)
        if key in log_karyawan_harian: # Ensure the key exists (should always if pre-populated correctly)
            log_karyawan_harian[key]['Data Log Mentah'].append(jam_obj)


    # Process aggregated logs for each employee per day (including those with no logs)
    df_hasil_list = []
    for key, data in sorted(log_karyawan_harian.items()): # Sort by key (Nama, Tanggal) for consistent output
        all_times = sorted(data['Data Log Mentah']) # Sort all times for the day

        if all_times: # Only calculate if there are actual logs for the day
            data['Jam Datang'] = all_times[0] # First log is Jam Datang
            data['Jam Pulang'] = all_times[-1] # Last log is Jam Pulang
            
            # Format 'Data Log Mentah' for display (as seen in screenshot's 'data' column)
            data['Data Log Mentah'] = '\n'.join([t.strftime('%H:%M') for t in all_times])

            # Calculate Durasi Jam Kerja
            if data['Jam Datang'] and data['Jam Pulang']:
                dt_datang = datetime.combine(data['Tanggal'], data['Jam Datang'])
                dt_pulang = datetime.combine(data['Tanggal'], data['Jam Pulang'])
                durasi_kerja_td = dt_pulang - dt_datang

                total_seconds = durasi_kerja_td.total_seconds()
                
                # Convert total seconds to HH:MM format as seen in screenshot
                hours = int(total_seconds // 3600)
                minutes = int((total_seconds % 3600) // 60)
                data['Durasi Jam Kerja'] = f"{hours:02}:{minutes:02}"
            else:
                data['Durasi Jam Kerja'] = "" # Set to empty string if no valid times for duration
        else:
            # If no logs for the day, set these fields to empty strings or None as required
            data['Jam Datang'] = ""
            data['Jam Pulang'] = ""
            data['Durasi Jam Kerja'] = ""
            data['Data Log Mentah'] = "" # Ensure this is also empty for blank days

        # Create a dictionary for the current row based on final columns
        row_data = {col: data.get(col) for col in kolom_final}
        df_hasil_list.append(row_data)

    df_hasil = pd.DataFrame(df_hasil_list)
    
    # Sort the final DataFrame for consistent output and grouping for Excel export
    df_hasil.sort_values(by=['Nama', 'Tanggal'], inplace=True)

    st.success("Pengolahan data selesai.")
    st.subheader("Data Hasil Akhir")
    st.dataframe(df_hasil) # Display the processed data

    return df_hasil, bulan_laporan, tahun_laporan


# Streamlit UI for file upload
uploaded_file = st.file_uploader("Pilih file Excel (.xlsx atau .xls)", type=["xlsx", "xls"])

if uploaded_file is not None:
    st.subheader("File yang Diunggah:")
    st.write(uploaded_file.name)

    df_processed, bulan_laporan_val, tahun_laporan_val = process_attendance_log(uploaded_file)

    if df_processed is not None:
        st.subheader("Download Hasil Pengolahan")

        output_buffer = io.BytesIO()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Rekap Absensi'

        # Removed page setup for A4 landscape as per user request
        # ws.page_setup.paperSize = PAPERSIZE.A4
        # ws.page_setup.orientation = ORIENTATION.LANDSCAPE
        # ws.page_setup.leftMargin = 0.5
        # ws.page_setup.rightMargin = 0.5
        # ws.page_setup.topMargin = 0.5
        # ws.page_setup.bottomMargin = 0.5


        # Define the exact column order for the Excel output's headers
        excel_display_headers = ['tanggal', 'data', 'datang', 'pulang', 'jumlah jam kerja']
        
        # Create a mapping from internal Python DataFrame names to desired Excel display names
        column_rename_map = {
            'Tanggal': 'tanggal',
            'Data Log Mentah': 'data',
            'Jam Datang': 'datang',
            'Jam Pulang': 'pulang',
            'Durasi Jam Kerja': 'jumlah jam kerja'
        }

        # Rename columns of df_processed to match the desired Excel display names
        df_excel_export_temp = df_processed.rename(columns=column_rename_map)
        
        # Ensure we only have the desired columns in the correct order for iteration
        df_excel_export_temp = df_excel_export_temp[['Nama'] + excel_display_headers]

        # Get unique employee names
        unique_employees = df_excel_export_temp['Nama'].unique().tolist()
        
        current_row = 1 # Start writing from row 1 in Excel

        # Define column offsets for the two tables
        # Table 1: starts from column 1 (A)
        # Table 2: starts from column 7 (G)
        # So, the offset for table 2 is 6 columns (G is 7th column, 1-indexed, so 6 from 1st col)
        col_offset_table2 = 6 

        # Define styles
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Light gray fill

        # Group employees into pairs for side-by-side display
        for i in range(0, len(unique_employees), 2):
            employee1 = unique_employees[i]
            employee2 = unique_employees[i+1] if i+1 < len(unique_employees) else None

            # Get data for employee 1
            df_emp1 = df_excel_export_temp[df_excel_export_temp['Nama'] == employee1].copy()
            df_emp1_data = df_emp1[excel_display_headers].values.tolist()

            # Calculate total duration for employee 1
            total_minutes_emp1 = sum(time_to_minutes(d) for d in df_emp1['jumlah jam kerja'].tolist())
            total_duration_emp1_hhmm = minutes_to_hhmm(total_minutes_emp1)

            # --- Write Employee 1's table ---
            # Employee Name Header
            ws.cell(row=current_row, column=1, value=employee1).font = bold_font
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
            # Column Headers
            for c_idx, header in enumerate(excel_display_headers):
                cell = ws.cell(row=current_row + 1, column=c_idx + 1, value=header) # +1 because data starts from column B for this table
                cell.font = bold_font
                cell.alignment = Alignment(horizontal='center')
                cell.fill = header_fill
                cell.border = thin_border

            # Data rows
            for r_data_idx, row_data in enumerate(df_emp1_data):
                for c_data_idx, cell_value in enumerate(row_data):
                    cell = ws.cell(row=current_row + 2 + r_data_idx, column=c_data_idx + 1, value=cell_value)
                    cell.border = thin_border
                    if c_data_idx == 0: # tanggal column
                        if isinstance(cell_value, date):
                            cell.value = cell_value.day # Just the day number
                        else: # Handle empty date cells for formatting
                            cell.value = ""
                        cell.number_format = '0' # Ensure it's treated as a number format
                        cell.alignment = Alignment(horizontal='center')
                    elif c_data_idx == 1: # data column (raw logs)
                        cell.alignment = Alignment(wrapText=True, vertical='top', horizontal='left')
                    elif c_data_idx in [2, 3, 4]: # datang, pulang, jumlah jam kerja
                        cell.alignment = Alignment(horizontal='center')
            
            # Total Row for Employee 1
            total_row_idx_emp1 = current_row + 2 + len(df_emp1_data)
            ws.cell(row=total_row_idx_emp1, column=excel_display_headers.index('jumlah jam kerja') + 1, value=total_duration_emp1_hhmm).font = bold_font
            ws.cell(row=total_row_idx_emp1, column=excel_display_headers.index('jumlah jam kerja') + 1).alignment = Alignment(horizontal='center')
            ws.cell(row=total_row_idx_emp1, column=excel_display_headers.index('jumlah jam kerja') + 1).border = thin_border
            ws.cell(row=total_row_idx_emp1, column=excel_display_headers.index('jumlah jam kerja') + 1).fill = header_fill # Apply fill to total cell
            ws.cell(row=total_row_idx_emp1, column=excel_display_headers.index('jumlah jam kerja') + 1).value = total_duration_emp1_hhmm


            # --- Write Employee 2's table (if exists) ---
            if employee2:
                # Get data for employee 2
                df_emp2 = df_excel_export_temp[df_excel_export_temp['Nama'] == employee2].copy()
                df_emp2_data = df_emp2[excel_display_headers].values.tolist()
                total_minutes_emp2 = sum(time_to_minutes(d) for d in df_emp2['jumlah jam kerja'].tolist())
                total_duration_emp2_hhmm = minutes_to_hhmm(total_minutes_emp2)

                # Employee Name Header
                ws.cell(row=current_row, column=1 + col_offset_table2, value=employee2).font = bold_font
                ws.cell(row=current_row, column=1 + col_offset_table2).alignment = Alignment(horizontal='left')
                # Column Headers
                for c_idx, header in enumerate(excel_display_headers):
                    cell = ws.cell(row=current_row + 1, column=c_idx + 1 + col_offset_table2, value=header)
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = header_fill
                    cell.border = thin_border

                # Data rows
                for r_data_idx, row_data in enumerate(df_emp2_data):
                    for c_data_idx, cell_value in enumerate(row_data):
                        cell = ws.cell(row=current_row + 2 + r_data_idx, column=c_data_idx + 1 + col_offset_table2, value=cell_value)
                        cell.border = thin_border
                        if c_data_idx == 0: # tanggal column
                            if isinstance(cell_value, date):
                                cell.value = cell_value.day # Just the day number
                            else: # Handle empty date cells for formatting
                                cell.value = ""
                            cell.number_format = '0' # Ensure it's treated as a number format
                            cell.alignment = Alignment(horizontal='center')
                        elif c_data_idx == 1: # data column (raw logs)
                            cell.alignment = Alignment(wrapText=True, vertical='top', horizontal='left')
                        elif c_data_idx in [2, 3, 4]: # datang, pulang, jumlah jam kerja
                            cell.alignment = Alignment(horizontal='center')
                
                # Total Row for Employee 2
                total_row_idx_emp2 = current_row + 2 + len(df_emp2_data)
                ws.cell(row=total_row_idx_emp2, column=excel_display_headers.index('jumlah jam kerja') + 1 + col_offset_table2, value=total_duration_emp2_hhmm).font = bold_font
                ws.cell(row=total_row_idx_emp2, column=excel_display_headers.index('jumlah jam kerja') + 1 + col_offset_table2).alignment = Alignment(horizontal='center')
                ws.cell(row=total_row_idx_emp2, column=excel_display_headers.index('jumlah jam kerja') + 1 + col_offset_table2).border = thin_border
                ws.cell(row=total_row_idx_emp2, column=excel_display_headers.index('jumlah jam kerja') + 1 + col_offset_table2).fill = header_fill # Apply fill to total cell
                ws.cell(row=total_row_idx_emp2, column=excel_display_headers.index('jumlah jam kerja') + 1 + col_offset_table2).value = total_duration_emp2_hhmm


            # Determine the maximum height of the two tables (including name, headers, data, and total row)
            # 1 for name row, 1 for header row, max_rows_in_pair for data rows, 1 for total row, 2 for blank lines
            max_rows_in_pair = max(len(df_emp1_data), len(df_emp2_data) if employee2 else 0)
            current_row += (1 + 1 + max_rows_in_pair + 1 + 2) # Name + Headers + Data + Total + 2 blank lines for separation

        # Set column widths for all relevant columns
        # Columns for table 1 (starts at column 1)
        ws.column_dimensions['A'].width = 10 # For employee name in column A (or blank separator)
        for i, col_name in enumerate(excel_display_headers):
            column_letter = chr(ord('A') + i + 1) # B, C, D, E, F
            if col_name == 'data':
                 ws.column_dimensions[column_letter].width = 15 # Adjust width for data column
            else:
                 ws.column_dimensions[column_letter].width = 12 # Default width for other columns

        # Columns for table 2 (starts at column 7)
        ws.column_dimensions['G'].width = 10 # For employee name in column G (or blank separator)
        for i, col_name in enumerate(excel_display_headers):
            column_letter = chr(ord('A') + i + 1 + col_offset_table2) # H, I, J, K, L
            if col_name == 'data':
                 ws.column_dimensions[column_letter].width = 15 # Adjust width for data column
            else:
                 ws.column_dimensions[column_letter].width = 12 # Default width for other columns


        wb.save(output_buffer)
        output_buffer.seek(0)

        st.download_button(
            label="Unduh File Excel Hasil",
            data=output_buffer.getvalue(),
            file_name='Rekap_Absensi_Per_Nama.xlsx',
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.success("File Excel berhasil dibuat dengan format per karyawan dan baris pemisah antar karyawan.")

