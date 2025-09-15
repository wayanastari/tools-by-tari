import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from datetime import datetime

def copy_cell_style(source_cell, target_cell):
    """Copy style from source cell to target cell"""
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            color=source_cell.font.color
        )
    if source_cell.border:
        target_cell.border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom
        )
    if source_cell.fill:
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )
    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            wrap_text=source_cell.alignment.wrap_text
        )

def find_form_boundaries(template_ws):
    """Find the boundaries of the form template"""
    max_row = 0
    max_col = 0
    
    for row in template_ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    
    return max_row, max_col

def map_data_to_form(data_row, template_ws, output_ws, start_row):
    """Map data from a row to the form template"""
    
    # Mapping dictionary - adjust these based on your exact column names
    field_mapping = {
        'Nama Bayi/Balita': 'Nama Bayi/Balita',
        'NIK': 'NIK',
        'Tanggal Lahir': 'Tanggal Lahir',
        'BB': 'Berat Badan Lahir',
        'TB': 'Panjang Badan Lahir',
        'AYAH': 'Nama Ayah/Ibu',
        'IBU': 'Nama Ayah/Ibu',
        'Alamat': 'Alamat',
        'No. Hp': 'No. Hp'
    }
    
    # Get form boundaries
    form_height, form_width = find_form_boundaries(template_ws)
    
    # Copy template structure
    for row_idx in range(1, form_height + 1):
        for col_idx in range(1, form_width + 1):
            source_cell = template_ws.cell(row=row_idx, column=col_idx)
            target_cell = output_ws.cell(row=start_row + row_idx - 1, column=col_idx)
            
            # Copy value and style
            target_cell.value = source_cell.value
            copy_cell_style(source_cell, target_cell)
    
    # Fill in the data
    for row_idx in range(1, form_height + 1):
        for col_idx in range(1, form_width + 1):
            source_cell = template_ws.cell(row=row_idx, column=col_idx)
            target_cell = output_ws.cell(row=start_row + row_idx - 1, column=col_idx)
            
            if source_cell.value:
                cell_text = str(source_cell.value)
                
                # Check if this cell contains a field we need to fill
                if 'Nama Bayi/Balita' in cell_text and 'Nama Bayi/Balita' in data_row:
                    # Find the cell next to it (usually contains dots) to fill with data
                    next_cell = output_ws.cell(row=start_row + row_idx - 1, column=col_idx + 1)
                    next_cell.value = data_row['Nama Bayi/Balita']
                
                elif 'NIK' in cell_text and 'NIK' in data_row:
                    next_cell = output_ws.cell(row=start_row + row_idx - 1, column=col_idx + 1)
                    next_cell.value = data_row['NIK']
                
                elif 'Tanggal Lahir' in cell_text and 'TANGGAL LAHIR' in data_row:
                    next_cell = output_ws.cell(row=start_row + row_idx - 1, column=col_idx + 1)
                    # Format date if needed
                    date_val = data_row['TANGGAL LAHIR']
                    if pd.notna(date_val):
                        if isinstance(date_val, str):
                            next_cell.value = date_val
                        else:
                            next_cell.value = date_val.strftime('%d-%m-%Y') if hasattr(date_val, 'strftime') else str(date_val)
                
                elif 'Berat Badan Lahir' in cell_text and 'BB' in data_row:
                    next_cell = output_ws.cell(row=start_row + row_idx - 1, column=col_idx + 1)
                    next_cell.value = f"{data_row['BB']} Kg" if pd.notna(data_row['BB']) else ""
                
                elif 'Panjang Badan Lahir' in cell_text and 'TB' in data_row:
                    next_cell = output_ws.cell(row=start_row + row_idx - 1, column=col_idx + 1)
                    next_cell.value = f"{data_row['TB']} Cm" if pd.notna(data_row['TB']) else ""
                
                elif 'Nama Ayah/Ibu' in cell_text:
                    next_cell = output_ws.cell(row=start_row + row_idx - 1, column=col_idx + 1)
                    ayah = data_row.get('AYAH', '')
                    ibu = data_row.get('IBU', '')
                    next_cell.value = f"Ayah: {ayah}, Ibu: {ibu}"
                
                # Add more mappings as needed based on your form structure
    
    return form_height

def generate_forms(data_df, template_file):
    """Generate multiple forms from data"""
    
    # Load template
    template_wb = load_workbook(template_file)
    template_ws = template_wb.active
    
    # Create output workbook
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "Generated Forms"
    
    # Get form dimensions
    form_height, form_width = find_form_boundaries(template_ws)
    
    current_row = 1
    spacing_between_forms = 3  # Extra rows between forms for printing
    
    # Generate form for each data row
    for index, data_row in data_df.iterrows():
        form_height_used = map_data_to_form(data_row, template_ws, output_ws, current_row)
        current_row += form_height_used + spacing_between_forms
        
        # Add page break suggestion (you can adjust this)
        if (index + 1) % 2 == 0:  # Every 2 forms, add extra spacing
            current_row += 2
    
    # Save to BytesIO
    output_file = io.BytesIO()
    output_wb.save(output_file)
    output_file.seek(0)
    
    return output_file

# Streamlit App
def main():
    st.set_page_config(
        page_title="Excel Form Generator", 
        page_icon="📋",
        layout="wide"
    )
    
    st.title("📋 Excel Form Generator")
    st.markdown("Upload your data file and form template to generate multiple forms automatically!")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("1. Upload Data File")
        data_file = st.file_uploader(
            "Choose your data Excel file", 
            type=['xlsx', 'xls'],
            help="File containing the records to be processed"
        )
        
        if data_file:
            try:
                data_df = pd.read_excel(data_file)
                st.success(f"✅ Data loaded: {len(data_df)} records found")
                
                # Show preview
                st.subheader("Data Preview")
                st.dataframe(data_df.head(), use_container_width=True)
                
                # Show column info
                st.subheader("Available Columns")
                st.write(", ".join(data_df.columns.tolist()))
                
            except Exception as e:
                st.error(f"Error reading data file: {str(e)}")
                data_df = None
    
    with col2:
        st.subheader("2. Upload Form Template")
        template_file = st.file_uploader(
            "Choose your form template Excel file", 
            type=['xlsx', 'xls'],
            help="Template form that will be filled with data"
        )
        
        if template_file:
            try:
                template_df = pd.read_excel(template_file)
                st.success("✅ Template loaded successfully")
                
                # Show template preview
                st.subheader("Template Preview")
                st.dataframe(template_df.head(), use_container_width=True)
                
            except Exception as e:
                st.error(f"Error reading template file: {str(e)}")
    
    # Generate forms section
    if data_file and template_file:
        st.markdown("---")
        st.subheader("3. Generate Forms")
        
        col3, col4, col5 = st.columns([1, 2, 1])
        
        with col4:
            if st.button("🚀 Generate Forms", type="primary", use_container_width=True):
                try:
                    with st.spinner("Generating forms... Please wait."):
                        output_file = generate_forms(data_df, template_file)
                    
                    st.success(f"✅ Successfully generated {len(data_df)} forms!")
                    
                    # Download button
                    st.download_button(
                        label="📥 Download Generated Forms",
                        data=output_file.getvalue(),
                        file_name=f"Generated_Forms_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                except Exception as e:
                    st.error(f"Error generating forms: {str(e)}")
                    st.write("Please check that your data columns match the expected format.")
    
    # Instructions
    with st.expander("📖 Instructions"):
        st.markdown("""
        ### How to use:
        
        1. **Prepare your data file**: Excel file with records containing fields like:
           - Nama Bayi/Balita
           - NIK
           - TANGGAL LAHIR
           - BB (Berat Badan)
           - TB (Tinggi Badan)
           - AYAH, IBU
           - etc.
        
        2. **Prepare your form template**: Excel file with the form layout you want to use
        
        3. **Upload both files** using the file uploaders above
        
        4. **Click Generate Forms** to create multiple forms
        
        5. **Download the result** - ready for printing!
        
        ### Notes:
        - Forms will be arranged vertically for easy printing
        - Spacing between forms is optimized for A4 paper
        - Original formatting from template will be preserved
        """)

if __name__ == "__main__":
    main()