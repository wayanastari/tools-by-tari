import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
import re

def copy_cell_style(source_cell, target_cell):
    """Copy style from source cell to target cell"""
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            color=source_cell.font.color
        )
    if source_cell.border:
        target_cell.border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom
        )
    if source_cell.fill:
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )
    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            wrap_text=source_cell.alignment.wrap_text
        )

def copy_row_heights_and_col_widths(source_ws, target_ws, source_start_row, target_start_row, num_rows):
    """Copy row heights and column widths from source to target"""
    # Copy row heights
    for i in range(num_rows):
        source_row = source_start_row + i
        target_row = target_start_row + i
        if source_ws.row_dimensions[source_row].height:
            target_ws.row_dimensions[target_row].height = source_ws.row_dimensions[source_row].height
    
    # Copy column widths
    for col in range(1, source_ws.max_column + 1):
        col_letter = get_column_letter(col)
        if source_ws.column_dimensions[col_letter].width:
            target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width

def find_form_boundaries(template_ws):
    """Find the boundaries of the form template"""
    max_row = 0
    max_col = 0
    
    for row in template_ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    
    return max_row, max_col

def format_date_value(date_val):
    """Format date value to string"""
    if pd.isna(date_val) or date_val is None:
        return ""
    
    if isinstance(date_val, str):
        return date_val
    
    # If it's a datetime object
    if hasattr(date_val, 'strftime'):
        return date_val.strftime('%d-%m-%Y')
    
    return str(date_val)

def find_dots_column(template_ws, row_idx):
    """Find the column that contains dots (:) for data entry"""
    for col_idx in range(1, template_ws.max_column + 1):
        cell = template_ws.cell(row=row_idx, column=col_idx)
        if cell.value and ':' in str(cell.value):
            return col_idx
    return None

def map_data_to_form(data_row, template_ws, output_ws, start_row):
    """Map data from a row to the form template"""
    
    # Get form boundaries
    form_height, form_width = find_form_boundaries(template_ws)
    
    # First, copy the entire template structure
    for row_idx in range(1, form_height + 1):
        for col_idx in range(1, form_width + 1):
            source_cell = template_ws.cell(row=row_idx, column=col_idx)
            target_cell = output_ws.cell(row=start_row + row_idx - 1, column=col_idx)
            
            # Copy value and style
            target_cell.value = source_cell.value
            copy_cell_style(source_cell, target_cell)
    
    # Copy row heights and column widths
    copy_row_heights_and_col_widths(template_ws, output_ws, 1, start_row, form_height)
    
    # Track if we've found Ayah/Ibu section
    ayah_row = None
    ibu_row = None
    
    # Now fill in the actual data
    for row_idx in range(1, form_height + 1):
        # Get the cell in column A (labels)
        label_cell = template_ws.cell(row=row_idx, column=1)
        
        if label_cell.value:
            label_text = str(label_cell.value).lower().strip()
            
            # Find where the dots (:) are for data entry
            dots_col = find_dots_column(template_ws, row_idx)
            if dots_col is None:
                continue
                
            target_data_cell = output_ws.cell(row=start_row + row_idx - 1, column=dots_col)
            
            # Map data based on label
            if 'nama bayi' in label_text or 'nama anak' in label_text:
                if 'Nama Bayi/Balita' in data_row.index:
                    target_data_cell.value = data_row['Nama Bayi/Balita']
                elif 'NAMA ANAK' in data_row.index:
                    target_data_cell.value = data_row['NAMA ANAK']
            
            elif 'nik' in label_text:
                if 'NIK' in data_row.index:
                    nik_value = data_row['NIK']
                    # Handle NIK as string to prevent scientific notation
                    if pd.notna(nik_value):
                        target_data_cell.value = str(nik_value).replace('.0', '')
                        # Set cell format to text
                        target_data_cell.number_format = '@'
            
            elif 'tanggal' in label_text and 'lahir' in label_text:
                if 'TANGGAL LAHIR' in data_row.index:
                    target_data_cell.value = format_date_value(data_row['TANGGAL LAHIR'])
            
            elif 'berat badan' in label_text:
                if 'BB' in data_row.index:
                    bb_value = data_row['BB']
                    if pd.notna(bb_value):
                        target_data_cell.value = str(bb_value)
            
            elif 'panjang badan' in label_text or 'tinggi badan' in label_text:
                if 'TB' in data_row.index:
                    tb_value = data_row['TB']
                    if pd.notna(tb_value):
                        target_data_cell.value = str(tb_value)
            
            elif 'nama ayah' in label_text and 'ibu' not in label_text:
                # This is specifically for Ayah row
                ayah_row = row_idx
                if 'AYAH' in data_row.index:
                    ayah_value = data_row['AYAH']
                    if pd.notna(ayah_value):
                        target_data_cell.value = str(ayah_value)
            
            elif 'nama' in label_text and 'ibu' in label_text and 'ayah' not in label_text:
                # This is specifically for Ibu row
                ibu_row = row_idx
                if 'IBU' in data_row.index:
                    ibu_value = data_row['IBU']
                    if pd.notna(ibu_value):
                        target_data_cell.value = str(ibu_value)
            
            elif 'ayah' in label_text and 'ibu' in label_text:
                # Combined Ayah/Ibu field - use first one found
                ayah = data_row.get('AYAH', '') if 'AYAH' in data_row.index else ''
                if pd.notna(ayah) and ayah:
                    target_data_cell.value = str(ayah)
                    # Check if next row is for Ibu
                    next_row_idx = row_idx + 1
                    if next_row_idx <= form_height:
                        next_dots_col = find_dots_column(template_ws, next_row_idx)
                        if next_dots_col:
                            next_target_cell = output_ws.cell(row=start_row + next_row_idx - 1, column=next_dots_col)
                            ibu = data_row.get('IBU', '') if 'IBU' in data_row.index else ''
                            if pd.notna(ibu) and ibu:
                                next_target_cell.value = str(ibu)
            
            elif 'alamat' in label_text:
                if 'ALAMAT' in data_row.index:
                    alamat_value = data_row['ALAMAT']
                    if pd.notna(alamat_value):
                        target_data_cell.value = str(alamat_value)
            
            elif 'no' in label_text and 'hp' in label_text:
                if 'No. Hp' in data_row.index:
                    hp_value = data_row['No. Hp']
                    if pd.notna(hp_value):
                        target_data_cell.value = str(hp_value)
    
    # Handle gender selection (Laki-laki/Perempuan)
    if 'JENIS KELAMIN' in data_row.index or 'JENIS' in data_row.index:
        gender_col = 'JENIS KELAMIN' if 'JENIS KELAMIN' in data_row.index else 'JENIS'
        gender_value = data_row[gender_col]
        
        # Find the cell containing "(Laki-laki/Perempuan)" and mark the appropriate choice
        for row_idx in range(1, form_height + 1):
            for col_idx in range(1, form_width + 1):
                source_cell = template_ws.cell(row=row_idx, column=col_idx)
                if source_cell.value and 'laki' in str(source_cell.value).lower() and 'perempuan' in str(source_cell.value).lower():
                    target_cell = output_ws.cell(row=start_row + row_idx - 1, column=col_idx)
                    
                    if pd.notna(gender_value):
                        if str(gender_value).upper() == 'L' or 'laki' in str(gender_value).lower():
                            target_cell.value = "(Laki-laki/Perempuan) ✓ Laki-laki"
                        elif str(gender_value).upper() == 'P' or 'perempuan' in str(gender_value).lower():
                            target_cell.value = "(Laki-laki/Perempuan) ✓ Perempuan"
                    break
    
    return form_height

def generate_forms(data_df, template_file):
    """Generate multiple forms from data"""
    
    # Load template
    template_wb = load_workbook(template_file)
    template_ws = template_wb.active
    
    # Create output workbook
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "Generated Forms"
    
    # Get form dimensions
    form_height, form_width = find_form_boundaries(template_ws)
    
    current_row = 1
    spacing_between_forms = 5  # Extra rows between forms for printing
    
    # Generate form for each data row
    for index, data_row in data_df.iterrows():
        st.write(f"Processing record {index + 1}/{len(data_df)}: {data_row.get('Nama Bayi/Balita', 'Unknown')}")
        
        form_height_used = map_data_to_form(data_row, template_ws, output_ws, current_row)
        current_row += form_height_used + spacing_between_forms
        
        # Add extra spacing every few forms for page breaks
        if (index + 1) % 3 == 0:  # Every 3 forms, add extra spacing
            current_row += 3
    
    # Auto-adjust column widths if not set
    for col in range(1, form_width + 1):
        col_letter = get_column_letter(col)
        if not output_ws.column_dimensions[col_letter].width:
            max_length = 0
            for row in range(1, current_row):
                cell_value = output_ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set reasonable width
            if max_length > 0:
                output_ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Save to BytesIO
    output_file = io.BytesIO()
    output_wb.save(output_file)
    output_file.seek(0)
    
    return output_file

# Streamlit App
def main():
    st.set_page_config(
        page_title="Excel Form Generator", 
        page_icon="📋",
        layout="wide"
    )
    
    st.title("📋 Excel Form Generator")
    st.markdown("Upload your data file and form template to generate multiple forms automatically!")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("1. Upload Data File")
        data_file = st.file_uploader(
            "Choose your data Excel file", 
            type=['xlsx', 'xls'],
            help="File containing the records to be processed"
        )
        
        if data_file:
            try:
                # Try different sheet names or use first sheet
                data_df = pd.read_excel(data_file)
                st.success(f"✅ Data loaded: {len(data_df)} records found")
                
                # Show preview
                st.subheader("Data Preview")
                st.dataframe(data_df.head(), use_container_width=True)
                
                # Show column info
                st.subheader("Available Columns")
                cols_text = ", ".join([f"'{col}'" for col in data_df.columns.tolist()])
                st.text(cols_text)
                
            except Exception as e:
                st.error(f"Error reading data file: {str(e)}")
                data_df = None
    
    with col2:
        st.subheader("2. Upload Form Template")
        template_file = st.file_uploader(
            "Choose your form template Excel file", 
            type=['xlsx', 'xls'],
            help="Template form that will be filled with data"
        )
        
        if template_file:
            st.success("✅ Template loaded successfully")
    
    # Generate forms section
    if data_file and template_file:
        st.markdown("---")
        st.subheader("3. Generate Forms")
        
        # Options
        st.subheader("Options")
        col3, col4 = st.columns(2)
        
        with col3:
            spacing = st.number_input("Spacing between forms (rows)", min_value=2, max_value=10, value=5)
            
        with col4:
            forms_per_page = st.number_input("Forms per page suggestion", min_value=1, max_value=5, value=3)
        
        if st.button("🚀 Generate Forms", type="primary", use_container_width=True):
            try:
                with st.spinner("Generating forms... Please wait."):
                    progress_bar = st.progress(0)
                    
                    # Update progress during generation
                    output_file = generate_forms(data_df, template_file)
                    progress_bar.progress(100)
                
                st.success(f"✅ Successfully generated {len(data_df)} forms!")
                
                # Download button
                st.download_button(
                    label="📥 Download Generated Forms",
                    data=output_file.getvalue(),
                    file_name=f"Generated_Forms_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
            except Exception as e:
                st.error(f"Error generating forms: {str(e)}")
                st.write("Detailed error information:")
                st.exception(e)
    
    # Instructions
    with st.expander("📖 Instructions & Tips"):
        st.markdown("""
        ### How to use:
        
        1. **Prepare your data file**: Excel file with records containing fields like:
           - Nama Bayi/Balita atau NAMA ANAK
           - NIK
           - TANGGAL LAHIR
           - BB (Berat Badan)
           - TB (Tinggi Badan)
           - AYAH, IBU
           - ALAMAT
           - No. Hp
        
        2. **Prepare your form template**: Excel file with form layout where:
           - Column A contains field labels (Nama Bayi, NIK, etc.)
           - Column B will be filled with data
           - Other columns can contain formatting or units
        
        3. **Upload both files** using the file uploaders above
        
        4. **Adjust settings** if needed (spacing, forms per page)
        
        5. **Click Generate Forms** to create multiple forms
        
        6. **Download the result** - ready for printing!
        
        ### Tips:
        - Forms will be arranged vertically with spacing for easy printing
        - Original formatting and styling will be preserved
        - NIK numbers will be formatted properly (no scientific notation)
        - Dates will be formatted as DD-MM-YYYY
        - The program automatically detects field labels in your template
        
        ### Troubleshooting:
        - Make sure column names in data match expected format
        - Check that template has clear field labels in column A
        - Verify data types (numbers, dates) are correctly formatted
        """)

if __name__ == "__main__":
    main()